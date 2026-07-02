from __future__ import annotations

import hashlib
import itertools
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from bricksmart.catalog.models import CatalogBlockDefinition, WorkbookCatalog
from bricksmart.exceptions import CatalogConfigurationError

_ID_ALIASES = (
    "block_family", "block_type", "block_id", "catalog_id", "block_name", "part_type", "part_id", "name"
)
_CATEGORY_ALIASES = ("category", "block_category", "block_family", "family", "type")
_ALLOWED_DIMENSION_ALIASES = (
    "geometry_size", "allowed_dimensions", "allowed_orientations", "allowed_sizes",
    "packing_dimensions", "rotated_dimensions", "orientation_dimensions",
)
_DIMENSION_X_ALIASES = ("size_x", "x_size", "dim_x", "studs_x", "length", "width")
_DIMENSION_Y_ALIASES = ("size_y", "y_size", "dim_y", "studs_y", "depth")
_DIMENSION_Z_ALIASES = ("size_z", "z_size", "dim_z", "studs_z", "height")
_STRUCTURAL_ALIASES = (
    "structural_eligible", "is_structural", "packing_eligible", "structural_block"
)
_PRIORITY_ALIASES = ("packing_priority", "planner_priority", "priority", "packing_rank")
_COLOR_ALIASES = (
    "display_color", "visualization_color", "render_color", "block_color",
    "hex_color", "color_hex", "colour", "color",
)
_ROTATION_ALIASES = (
    "allowed_rotations", "rotation_policy", "orientation_policy", "rotations", "rotation"
)
_MALE_FACE_ALIASES = (
    "primary_male_faces", "male_faces", "male_face", "male_connectors", "male_connector_faces"
)
_FEMALE_FACE_ALIASES = (
    "primary_female_faces", "female_faces", "female_face", "female_connectors", "female_connector_faces"
)

_DIMENSION_RE = re.compile(r"(?<!\d)(\d+)\s*[x×,]\s*(\d+)\s*[x×,]\s*(\d+)(?!\d)", re.IGNORECASE)
_TOKEN_SPLIT_RE = re.compile(r"[;,|\n]+")
_TRUE_VALUES = {"true", "yes", "y", "1", "enabled", "eligible"}
_FALSE_VALUES = {"false", "no", "n", "0", "disabled", "ineligible", ""}


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _first_present(row: Mapping[str, Any], aliases: Sequence[str]) -> Any:
    for alias in aliases:
        if alias in row and row[alias] not in (None, ""):
            return row[alias]
    return None


def _json_safe(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _parse_bool(value: Any, *, field: str, block_type: str) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return bool(value)
    normalized = str(value).strip().lower()
    if normalized in _TRUE_VALUES:
        return True
    if normalized in _FALSE_VALUES:
        return False
    raise CatalogConfigurationError(
        f"Cannot parse {field}={value!r} for catalog block {block_type}"
    )


def _parse_int(value: Any, *, default: int, field: str, block_type: str) -> int:
    if value in (None, ""):
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError) as exc:
        raise CatalogConfigurationError(
            f"Cannot parse {field}={value!r} for catalog block {block_type}"
        ) from exc


def _parse_face_tokens(value: Any) -> tuple[str, ...]:
    if value in (None, ""):
        return ()
    if isinstance(value, (list, tuple, set)):
        raw = [str(item).strip() for item in value]
    else:
        text = str(value).strip()
        if text.startswith("["):
            try:
                payload = json.loads(text)
            except json.JSONDecodeError:
                payload = None
            if isinstance(payload, list):
                raw = [str(item).strip() for item in payload]
            else:
                raw = [item.strip() for item in _TOKEN_SPLIT_RE.split(text)]
        else:
            raw = [item.strip() for item in _TOKEN_SPLIT_RE.split(text)]
    return tuple(dict.fromkeys(item for item in raw if item))


def _parse_dimension_value(value: Any) -> tuple[tuple[int, int, int], ...]:
    if value in (None, ""):
        return ()
    dimensions: list[tuple[int, int, int]] = []
    if isinstance(value, (list, tuple)):
        candidates = value
    else:
        text = str(value).strip()
        candidates: Any = None
        if text.startswith("["):
            try:
                candidates = json.loads(text)
            except json.JSONDecodeError:
                candidates = None
        if candidates is None:
            candidates = _DIMENSION_RE.findall(text)
    for item in candidates:
        if isinstance(item, str):
            matches = _DIMENSION_RE.findall(item)
            dimensions.extend(tuple(int(v) for v in match) for match in matches)
        elif isinstance(item, (list, tuple)) and len(item) == 3:
            dimensions.append(tuple(int(v) for v in item))
    return tuple(dict.fromkeys(dimensions))


def _rotation_tokens(value: Any) -> tuple[str, ...]:
    if value in (None, ""):
        return ()
    if isinstance(value, (list, tuple, set)):
        tokens = [str(item).strip().lower() for item in value]
    else:
        tokens = [item.strip().lower() for item in _TOKEN_SPLIT_RE.split(str(value))]
    return tuple(dict.fromkeys(token for token in tokens if token))


def _dimensions_from_xyz(row: Mapping[str, Any], *, block_type: str) -> tuple[int, int, int] | None:
    values = [
        _first_present(row, _DIMENSION_X_ALIASES),
        _first_present(row, _DIMENSION_Y_ALIASES),
        _first_present(row, _DIMENSION_Z_ALIASES),
    ]
    if all(value in (None, "") for value in values):
        return None
    if any(value in (None, "") for value in values):
        raise CatalogConfigurationError(
            f"Catalog block {block_type} has an incomplete X/Y/Z dimension triplet"
        )
    try:
        parsed = tuple(int(float(value)) for value in values)
    except (TypeError, ValueError) as exc:
        raise CatalogConfigurationError(
            f"Catalog block {block_type} contains non-integer dimensions: {values}"
        ) from exc
    if any(value <= 0 for value in parsed):
        raise CatalogConfigurationError(
            f"Catalog block {block_type} dimensions must all be positive: {parsed}"
        )
    return parsed  # type: ignore[return-value]


def _expand_dimensions_by_rotation(
    base: tuple[int, int, int] | None,
    rotations: tuple[str, ...],
) -> tuple[tuple[int, int, int], ...]:
    if base is None:
        return ()
    normalized = " ".join(rotations)
    if not normalized or normalized in {"none", "fixed", "0", "no_rotation"}:
        return (base,)
    if any(token in normalized for token in ("all", "any", "3d", "xyz", "permutation")):
        return tuple(dict.fromkeys(itertools.permutations(base, 3)))
    if any(token in normalized for token in ("z", "xy", "90", "270", "quarter_turn")):
        x, y, z = base
        return tuple(dict.fromkeys((base, (y, x, z))))
    return (base,)


def _find_header_row(worksheet: Any, *, scan_limit: int = 40) -> tuple[int, list[str]] | None:
    for row_index, values in enumerate(
        worksheet.iter_rows(min_row=1, max_row=scan_limit, values_only=True), start=1
    ):
        headers = [_normalize_header(value) for value in values]
        if any(alias in headers for alias in _ID_ALIASES):
            return row_index, headers
    return None


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_block_catalog(path: str | Path) -> WorkbookCatalog:
    """Load the original BrickSmart XLSX workbook without creating a shadow catalog."""
    path = Path(path)
    if not path.exists():
        raise CatalogConfigurationError(
            f"Original block catalog workbook not found: {path}. "
            "Place block_definitions.xlsx at this exact path or pass --catalog."
        )
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise CatalogConfigurationError(
            "BrickSmart accepts only the original Excel catalog (.xlsx/.xlsm); "
            f"shadow CSV/JSON/YAML catalogs are prohibited: {path}"
        )
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise CatalogConfigurationError(
            "Reading block_definitions.xlsx requires openpyxl>=3.1"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    definitions: list[CatalogBlockDefinition] = []
    header_rows: dict[str, int] = {}
    sheets_read: list[str] = []

    preferred_sheet = workbook["Block Definitions"] if "Block Definitions" in workbook.sheetnames else None
    worksheets = [preferred_sheet] if preferred_sheet is not None else list(workbook.worksheets)

    for worksheet in worksheets:
        detected = _find_header_row(worksheet)
        if detected is None:
            continue
        header_row, headers = detected
        id_columns = [alias for alias in _ID_ALIASES if alias in headers]
        if not id_columns:
            continue
        sheets_read.append(worksheet.title)
        header_rows[worksheet.title] = header_row
        id_column = id_columns[0]
        for source_row, values in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            row = {
                header: value
                for header, value in zip(headers, values)
                if header
            }
            block_type = str(row.get(id_column) or "").strip()
            if not block_type:
                continue
            category = str(_first_present(row, _CATEGORY_ALIASES) or "").strip()
            explicit_structural = _first_present(row, _STRUCTURAL_ALIASES)
            structural_eligible = (
                _parse_bool(
                    explicit_structural,
                    field="structural_eligible",
                    block_type=block_type,
                )
                if explicit_structural not in (None, "")
                else "structural" in category.lower()
            )
            rotations = _rotation_tokens(_first_present(row, _ROTATION_ALIASES))
            dimensions = _parse_dimension_value(
                _first_present(row, _ALLOWED_DIMENSION_ALIASES)
            )
            if not dimensions:
                dimensions = _expand_dimensions_by_rotation(
                    _dimensions_from_xyz(row, block_type=block_type), rotations
                )
            if structural_eligible and not dimensions:
                raise CatalogConfigurationError(
                    f"Structural catalog block {block_type} on "
                    f"{worksheet.title}!{source_row} has no usable dimensions"
                )
            color = str(_first_present(row, _COLOR_ALIASES) or "").strip()
            if structural_eligible and not color:
                raise CatalogConfigurationError(
                    f"Structural catalog block {block_type} on "
                    f"{worksheet.title}!{source_row} has no visualization color"
                )
            priority = _parse_int(
                _first_present(row, _PRIORITY_ALIASES),
                default=0,
                field="packing_priority",
                block_type=block_type,
            )
            definitions.append(
                CatalogBlockDefinition(
                    block_type=block_type,
                    category=category,
                    allowed_dimensions=dimensions,
                    structural_eligible=structural_eligible,
                    packing_priority=priority,
                    display_color=color,
                    male_faces=_parse_face_tokens(_first_present(row, _MALE_FACE_ALIASES)),
                    female_faces=_parse_face_tokens(_first_present(row, _FEMALE_FACE_ALIASES)),
                    allowed_rotations=rotations,
                    source_sheet=worksheet.title,
                    source_row=source_row,
                    raw_metadata={key: _json_safe(value) for key, value in row.items()},
                )
            )

    if not definitions:
        raise CatalogConfigurationError(
            f"No block definition table with a recognized ID column was found in {path}"
        )
    seen: dict[str, CatalogBlockDefinition] = {}
    for item in definitions:
        if item.block_type in seen:
            previous = seen[item.block_type]
            raise CatalogConfigurationError(
                f"Duplicate block ID {item.block_type!r} in "
                f"{previous.source_sheet}!{previous.source_row} and "
                f"{item.source_sheet}!{item.source_row}"
            )
        seen[item.block_type] = item

    return WorkbookCatalog(
        source_path=path.resolve(),
        source_sha256=_sha256(path),
        definitions=tuple(definitions),
        sheets_read=tuple(sheets_read),
        header_rows=header_rows,
    )


def load_catalog_block_ids(path: str | Path) -> set[str]:
    return load_block_catalog(path).block_ids


def validate_inventory_against_catalog(
    inventory_block_types: Iterable[str], catalog_block_ids: set[str]
) -> None:
    unknown = sorted(set(inventory_block_types) - set(catalog_block_ids))
    if unknown:
        raise CatalogConfigurationError(
            "Inventory contains block types missing from block_definitions.xlsx: "
            + ", ".join(unknown)
        )


def validate_used_block_colors(
    used_block_types: Iterable[str], catalog: WorkbookCatalog
) -> None:
    missing = sorted(
        block_type
        for block_type in set(used_block_types)
        if not catalog.by_type.get(block_type)
        or not catalog.by_type[block_type].display_color
    )
    if missing:
        raise CatalogConfigurationError(
            "Used block types have no catalog visualization color: " + ", ".join(missing)
        )
